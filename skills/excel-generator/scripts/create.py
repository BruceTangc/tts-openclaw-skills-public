#!/usr/bin/env python3
"""
Excel 创建脚本 - 结合 openpyxl 和 xlsxwriter 的优势
"""

import sys
import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_excel(file_path, data, sheet_name='Sheet1', style_config=None):
    """
    创建 Excel 文件
    
    Args:
        file_path: 输出文件路径
        data: 2D 数组数据
        sheet_name: 工作表名称
        style_config: 样式配置
    """
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 应用样式（如果是标题行）
                if row_idx == 1 and style_config:
                    apply_header_style(cell, style_config.get('header', {}))
                
                # 应用边框
                if style_config and style_config.get('border'):
                    apply_border(cell)
        
        # 自动调整列宽
        if style_config and style_config.get('auto_width'):
            auto_adjust_column_widths(ws)
        
        # 应用交替行样式
        if style_config and style_config.get('alternating_rows'):
            apply_alternating_rows(ws, style_config)
        
        # 保存文件
        wb.save(file_path)
        
        return {
            'success': True,
            'output_file': file_path,
            'sheet_count': 1,
            'row_count': len(data),
            'column_count': len(data[0]) if data else 0
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def apply_header_style(cell, header_config):
    """应用标题样式"""
    # 加粗
    if header_config.get('bold'):
        cell.font = Font(bold=True)
    
    # 背景色
    if header_config.get('bgColor'):
        cell.fill = PatternFill(
            start_color=header_config['bgColor'],
            end_color=header_config['bgColor'],
            fill_type='solid'
        )
    
    # 文字颜色
    if header_config.get('color'):
        cell.font = Font(color=header_config['color'], bold=cell.font.bold)
    
    # 字体大小
    if header_config.get('fontSize'):
        cell.font = Font(size=header_config['fontSize'], bold=cell.font.bold)
    
    # 对齐方式
    cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_border(cell):
    """应用边框"""
    thin_border = Side(style='thin', color='000000')
    cell.border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )

def apply_alternating_rows(ws, style_config):
    """应用交替行样式"""
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for row_idx in range(2, ws.max_row + 1):  # 从第 2 行开始（跳过标题）
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                    cell.fill = light_fill

def auto_adjust_column_widths(ws):
    """自动调整列宽"""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # 添加 2 个字符的缓冲
        adjusted_width = min(max_length + 2, 50)  # 最大 50 个字符
        ws.column_dimensions[column_letter].width = adjusted_width

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='创建 Excel 文件')
    parser.add_argument('--file', required=True, help='输出文件路径')
    parser.add_argument('--sheet', default='Sheet1', help='工作表名称')
    parser.add_argument('--data', required=True, help='JSON 格式的 2D 数组数据')
    parser.add_argument('--style', help='JSON 格式的样式配置')
    
    args = parser.parse_args()
    
    # 解析数据
    data = json.loads(args.data)
    
    # 解析样式（如果有）
    style_config = None
    if args.style:
        style_config = json.loads(args.style)
    
    # 创建 Excel
    result = create_excel(args.file, data, args.sheet, style_config)
    
    # 输出结果
    print(json.dumps(result, ensure_ascii=False))
