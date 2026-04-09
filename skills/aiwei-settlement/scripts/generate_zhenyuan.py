#!/usr/bin/env python3
"""
生成南通震源机械有限公司结算单
"""

import sys
import os
sys.path.insert(0, '/home/admin/.local/lib/python3.11/site-packages')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def extract_cutting_times_from_pdf(pdf_path):
    """从 PDF 第一页提取 Cutting time 数据"""
    import fitz
    import re
    
    doc = fitz.open(pdf_path)
    text = doc[0].get_text()
    lines = text.split('\n')
    
    # 从 Material data 获取厚度
    thickness = None
    in_material = False
    for line in lines:
        if 'Material data' in line:
            in_material = True
            continue
        if in_material and 'Plan data' in line:
            break
        if in_material:
            thickness_match = re.match(r'^(\d+\.?\d*)\s*mm$', line.strip())
            if thickness_match:
                thickness = float(thickness_match.group(1))
                break
    
    # 查找 Job data 部分的 Cutting time
    cutting_time = None
    for i, line in enumerate(lines):
        if 'Cutting time' in line:
            for j in range(i+1, min(i+15, len(lines))):
                time_match = re.search(r'(\d+\.?\d*)\s*min', lines[j])
                if time_match:
                    cutting_time = float(time_match.group(1))
                    break
            if cutting_time:
                break
    
    if cutting_time and thickness:
        return [(thickness, cutting_time)]
    return []


def extract_cutting_times_batch(pdf_dir):
    """从目录中所有 PDF 提取 Cutting time，并按厚度累加"""
    import glob
    
    pdf_files = glob.glob(os.path.join(pdf_dir, '*.pdf'))
    thickness_times = {}
    
    for pdf_path in pdf_files:
        result = extract_cutting_times_from_pdf(pdf_path)
        if result:
            thickness, cutting_time = result[0]
            if thickness in thickness_times:
                thickness_times[thickness] += cutting_time
            else:
                thickness_times[thickness] = cutting_time
    
    return thickness_times


def create_zhenyuan_settlement(output_file, contract_no, data_rows=None, pdf_path=None, pdf_dir=None, company_info=None):
    """创建震源机械结算单"""
    
    # 默认公司信息
    if company_info is None:
        company_info = {
            'name_cn': '南通震源机械有限公司',
            'name_en': 'Nantong source of metal cutting Co.Ltd.',
            'address': '地址：江苏省如皋市中山西路 398 号',
            'contact': '联系人：谢万宏   电话：18962729888   传真：86-0513-87872988',
            'web': 'Http://www.ntzhenyuan.com   E-mail：2995279201@qq.com',
            'email2': 'ntzy.laser@foxmail.com'
        }
    
    # 如果提供了 PDF 文件夹，批量提取并累加相同厚度
    if pdf_dir and not data_rows:
        thickness_times = extract_cutting_times_batch(pdf_dir)
        data_rows = []
        for thickness in sorted(thickness_times.keys()):
            cutting_time = thickness_times[thickness]
            data_rows.append({
                'date': '',
                'thickness': thickness,
                'theoretical_time': round(cutting_time, 2),
                'processing_fee': 6.33
            })
    
    # 如果提供了 PDF 路径，自动提取 Cutting time
    elif pdf_path and not data_rows:
        cutting_times = extract_cutting_times_from_pdf(pdf_path)
        data_rows = []
        for thickness, cutting_time in cutting_times:
            data_rows.append({
                'date': '',
                'thickness': thickness,
                'theoretical_time': cutting_time,
                'processing_fee': 6.33
            })
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = contract_no
    
    # 样式定义
    normal_font = Font(size=10, name='宋体')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    summary_font = Font(bold=True, size=10, name='宋体')
    summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # ========== 公司抬头（第 1-5 行）==========
    ws.cell(row=1, column=1, value=company_info['name_cn'])
    ws.merge_cells(f'A1:L1')
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, size=24, name='宋体')
    cell.alignment = center_align
    
    ws.cell(row=2, column=1, value=company_info['name_en'])
    ws.merge_cells(f'A2:L2')
    ws.cell(row=2, column=1).alignment = center_align
    
    ws.cell(row=3, column=1, value=company_info['address'])
    ws.merge_cells(f'A3:L3')
    ws.cell(row=3, column=1).alignment = center_align
    
    ws.cell(row=4, column=1, value=company_info['contact'])
    ws.merge_cells(f'A4:L4')
    ws.cell(row=4, column=1).alignment = center_align
    
    ws.cell(row=5, column=1, value=company_info['web'])
    ws.merge_cells(f'A5:L5')
    ws.cell(row=5, column=1).alignment = center_align
    
    ws.row_dimensions[6].height = 15
    
    # ========== 合同编号（第 7-11 行）==========
    contracts = [
        f'1.合同编号：{contract_no}',
        '2.合同编号：AW-26-0215',
        '3.合同编号：AW-26-0229',
        '4.合同编号：AW-26-0278',
        '5.合同编号：AW-26-0339'
    ]
    
    for i, contract in enumerate(contracts, 7):
        ws.cell(row=i, column=1, value=contract)
        ws.merge_cells(f'A{i}:L{i}')
        cell = ws.cell(row=i, column=1)
        cell.alignment = left_align
    
    ws.row_dimensions[12].height = 15
    
    # ========== 表头（第 13 行）==========
    headers = [
        '生产日期', '厚度 mm', '', '理论时间分钟', '理论时间\n总计分钟',
        '结算时间分钟', '加工费\n1 分钟', '加工费 总价（元）',
        '材料费', '打孔攻丝总价', '折弯总价', '原材料 + 打孔攻丝费 + 折弯\n总计元'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # 第 1-13 行添加边框
    for row in range(1, 14):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = thin_border
    
    # ========== 结算系数（第 14 行）==========
    for col in range(1, 13):
        cell = ws.cell(row=14, column=col)
        cell.border = thin_border
        cell.alignment = center_align
    ws.cell(row=14, column=6, value=0.1)
    ws.cell(row=14, column=6).number_format = '0.0'
    
    # ========== 数据行（第 15-24 行，固定 10 行）==========
    standard_thickness = [1, 1.5, 2, 2.5, 3, 4, 5, 6, 10, 12]
    
    for i, thickness in enumerate(standard_thickness):
        row = 15 + i
        row_data = None
        
        for dr in data_rows:
            if abs(dr['thickness'] - thickness) < 0.01:
                row_data = dr
                break
        
        if row_data:
            ws.cell(row=row, column=2, value=thickness)
            ws.cell(row=row, column=5, value=row_data.get('theoretical_time', 0))
            ws.cell(row=row, column=6, value=f'=E{row}+(E{row}*F$14)')
            ws.cell(row=row, column=7, value=row_data.get('processing_fee', 0))
            ws.cell(row=row, column=8, value=f'=G{row}*F{row}')
            ws.cell(row=row, column=12, value=f'=H{row}+I{row}+J{row}+K{row}')
        
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align
    
    # ========== 合计行（第 25 行）==========
    ws.cell(row=25, column=1, value='合计：')
    ws.cell(row=25, column=1).font = summary_font
    ws.cell(row=25, column=1).fill = summary_fill
    ws.cell(row=25, column=12, value='=SUM(L15:L24)')
    ws.cell(row=25, column=12).number_format = '0.00'
    ws.cell(row=25, column=12).font = summary_font
    
    for col in range(1, 13):
        cell = ws.cell(row=25, column=col)
        cell.border = thin_border
        cell.alignment = center_align
    
    # 设置列宽
    column_widths = {
        'A': 11.0, 'B': 13.0, 'C': 13.0, 'D': 13.0, 'E': 13.0,
        'F': 13.0, 'G': 13.0, 'H': 12.625, 'I': 13.0, 'J': 13.0,
        'K': 13.0, 'L': 12.625,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 设置行高（完全匹配原始文件）
    ws.row_dimensions[1].height = 31.5   # 公司名
    ws.row_dimensions[2].height = 20.25  # 公司英文名
    ws.row_dimensions[3].height = 15.0   # 地址
    ws.row_dimensions[4].height = 15.0   # 联系人
    ws.row_dimensions[5].height = 15.0   # 网址邮箱
    ws.row_dimensions[6].height = 15.0   # 空行
    for row in range(7, 12):
        ws.row_dimensions[row].height = 15.0  # 合同编号
    ws.row_dimensions[12].height = 14.25  # 空行
    ws.row_dimensions[13].height = 57.0   # 表头（自动换行）
    ws.row_dimensions[14].height = 15.0   # 结算系数
    for row in range(15, 25):
        ws.row_dimensions[row].height = 14.25  # 数据行
    ws.row_dimensions[25].height = 18.75  # 合计
    
    # 设置数字格式
    for row in range(15, 25):
        thickness = ws.cell(row=row, column=2).value
        if thickness and thickness == int(thickness):
            ws.cell(row=row, column=2).number_format = '0'
        else:
            ws.cell(row=row, column=2).number_format = '0.0'
        ws.cell(row=row, column=5).number_format = '0.00'
        ws.cell(row=row, column=6).number_format = '0.00'
        ws.cell(row=row, column=7).number_format = '0.00'
        ws.cell(row=row, column=8).number_format = '0.00'
        ws.cell(row=row, column=12).number_format = '0.00'
    
    wb.save(output_file)
    return output_file


if __name__ == '__main__':
    pdf_dir = '/tmp/extracted_aw253331/AW-25-3331'
    output = create_zhenyuan_settlement(
        '/tmp/AW-25-3331_震源结算单_测试.xlsx',
        'AW-25-3331',
        pdf_dir=pdf_dir
    )
    print(f'✅ 结算单已生成：{output}')
