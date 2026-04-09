#!/usr/bin/env python3
"""
PDF to Excel - 从 ByWork PDF 工单提取板材数据生成 Excel 费用计算表
"""

import sys
import os
import re
import json
import argparse
from datetime import datetime

try:
    import fitz  # PyMuPDF
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(json.dumps({
        'success': False,
        'error': f'缺少依赖：{e}',
        'hint': '请运行：pip install PyMuPDF openpyxl'
    }))
    sys.exit(1)


def extract_order_from_filename(filename):
    """从 PDF 文件名提取订单号"""
    match = re.match(r'(AW-\d+-\d+)', filename)
    if match:
        return match.group(1)
    return 'AW-25-3331'


def get_material_from_filename_and_thickness(filename, thickness):
    """根据 PDF 文件名和厚度判断材质"""
    if filename.startswith('sus-'):
        return 'SUS'
    elif filename.startswith('dxb-'):
        return '镀锌板'
    elif 'Mn' in filename:
        return 'Q345'
    else:
        if thickness <= 2:
            return 'Q235 冷板'
        else:
            return 'Q235'


def adjust_thickness(pdf_thickness, material):
    """根据材质调整实际厚度"""
    if material in ['Q235', 'Q235 冷板', 'Q345'] and pdf_thickness >= 3:
        return pdf_thickness - 0.25
    return pdf_thickness


def parse_dimension(dim_str):
    """解析尺寸字符串"""
    match = re.search(r'(\d+)\s*x\s*(\d+)', dim_str)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None


def read_material_data(text):
    """从 Material data 部分读取板材汇总信息"""
    material_summary = {}
    lines = text.split('\n')
    
    in_material = False
    current_thickness = None
    
    for i, line in enumerate(lines):
        if 'Material data' in line:
            in_material = True
            continue
        if in_material and 'Plan data' in line:
            break
        
        if in_material:
            thickness_match = re.match(r'(\d+\.?\d*)\s*mm', line.strip())
            if thickness_match:
                current_thickness = float(thickness_match.group(1))
                continue
            
            dim_match = re.search(r'(\d+)\s*x\s*(\d+)\s*mm', line)
            if dim_match and current_thickness is not None:
                width = int(dim_match.group(1))
                height = int(dim_match.group(2))
                key = f'{width}x{height}'
                
                used = 0
                if i + 1 < len(lines):
                    used_match = re.search(r'(\d+)\s*/\s*(\d+)', lines[i + 1])
                    if used_match:
                        used = int(used_match.group(1))
                
                material_summary[key] = {
                    'thickness': current_thickness,
                    'used': used
                }
    
    return material_summary


def read_plan_data(text):
    """从 Plan data 部分读取所有 Plan 信息"""
    plans = []
    lines = text.split('\n')
    
    in_plan = False
    plan_data_lines = []
    
    for line in lines:
        if 'Plan data' in line:
            in_plan = True
            continue
        if in_plan and 'Flat part data' in line:
            break
        if in_plan:
            if any(h in line for h in ['Plan No.', 'File name', 'Plan dimension', 'Sheet dimension', 
                                        'Cycles', 'Cutting time', 'Waste Number of pa']):
                continue
            if line.strip():
                plan_data_lines.append(line.strip())
    
    i = 0
    plan_no = 1
    while i + 6 < len(plan_data_lines):
        plan = {'plan_no': plan_no}
        
        sheet_line = plan_data_lines[i + 2]
        dim_match = re.search(r'(\d+)\s*x\s*(\d+)\s*mm', sheet_line)
        if dim_match:
            plan['sheet_w'] = int(dim_match.group(1))
            plan['sheet_h'] = int(dim_match.group(2))
        else:
            i += 7
            continue
        
        cycles_line = plan_data_lines[i + 3]
        if cycles_line.isdigit():
            plan['cycles'] = int(cycles_line)
        else:
            i += 7
            continue
        
        waste_line = plan_data_lines[i + 5]
        waste_match = re.search(r'(\d+\.?\d*)\s*%', waste_line)
        if waste_match:
            plan['waste'] = float(waste_match.group(1))
        
        parts_line = plan_data_lines[i + 6]
        if parts_line.isdigit():
            plan['parts'] = int(parts_line)
        
        plans.append(plan)
        plan_no += 1
        i += 7
    
    return plans


def validate_extraction(material_summary, plans, pdf_file):
    """检验提取的数据"""
    errors = []
    warnings = []
    
    total_cycles = sum(p.get('cycles', 0) for p in plans)
    total_used = sum(m['used'] for m in material_summary.values())
    
    if total_cycles != total_used and total_cycles > 0:
        warnings.append(f'数量不匹配：Plan Cycles={total_cycles}, Material Used={total_used}')
    
    for plan in plans:
        if plan.get('parts', 0) == 0:
            errors.append(f"Plan {plan.get('plan_no', '?')} 是空排版（Number of parts=0）")
    
    for plan in plans:
        sheet_w = plan.get('sheet_w')
        sheet_h = plan.get('sheet_h')
        if sheet_w and sheet_h:
            matched = False
            for mat_key in material_summary.keys():
                mat_w, mat_h = parse_dimension(mat_key)
                if mat_w and mat_h:
                    if sheet_w <= mat_w and sheet_h <= mat_h:
                        matched = True
                        break
            if not matched and material_summary:
                warnings.append(f"Plan 尺寸 {sheet_w}x{sheet_h} 无对应的 Material 尺寸")
    
    return errors, warnings


def process_pdf(pdf_path, order_no=None):
    """处理单个 PDF 文件"""
    pdf_file = os.path.basename(pdf_path)
    
    if not order_no:
        order_no = extract_order_from_filename(pdf_file)
    
    doc = fitz.open(pdf_path)
    
    plan_text = ''
    for page in doc:
        text = page.get_text()
        if 'Plan data' in text and 'Sheet dimension' in text:
            plan_text = text
            break
    
    if not plan_text:
        for page in doc:
            plan_text += page.get_text()
    
    material_summary = read_material_data(plan_text)
    plans = read_plan_data(plan_text)
    
    errors, warnings = validate_extraction(material_summary, plans, pdf_file)
    
    excel_rows = []
    for plan in plans:
        if plan.get('parts', 0) == 0:
            continue
        
        thickness = 0
        for mat_key, mat_val in material_summary.items():
            mat_w, mat_h = parse_dimension(mat_key)
            if mat_w and mat_h:
                if plan.get('sheet_w') <= mat_w and plan.get('sheet_h') <= mat_h:
                    thickness = mat_val['thickness']
                    break
        
        if thickness == 0 and 'Thickness' in plan_text:
            thickness_match = re.search(r'(\d+\.?\d*)\s*mm', plan_text)
            if thickness_match:
                thickness = float(thickness_match.group(1))
        
        material = get_material_from_filename_and_thickness(pdf_file, thickness)
        actual_thickness = adjust_thickness(thickness, material)
        
        excel_rows.append({
            'C': order_no,
            'D': material,
            'E': plan.get('sheet_w', 0),
            'F': plan.get('sheet_h', 0),
            'G': actual_thickness,
            'H': plan.get('cycles', 0),
            'O': plan.get('waste', 0) / 100
        })
    
    return {
        'rows': excel_rows,
        'errors': errors,
        'warnings': warnings,
        'stats': {
            'material_count': len(material_summary),
            'plan_count': len(plans)
        }
    }


def create_excel(all_rows, output_path, order_no='AW-25-3331'):
    """创建 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '板材费用统计'
    
    headers = ['序号', '日期', '订单号', '材质（Q235/Q345/SUS）', '长（mm）', '宽（mm）', 
               '厚（mm）', '数量', '重量（单重 Kg）', '合计重量（Kg）', '板材利用率', 
               '废料重量（Kg）', '单价（Kg）', '板材价格', '废料率%']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    today = datetime.now().strftime('%Y-%m-%d')
    for i, row_data in enumerate(all_rows, 2):
        ws.cell(row=i, column=2, value=today)
        ws.cell(row=i, column=3, value=row_data['C'])
        ws.cell(row=i, column=4, value=row_data['D'])
        ws.cell(row=i, column=5, value=row_data['E'])
        ws.cell(row=i, column=6, value=row_data['F'])
        ws.cell(row=i, column=7, value=row_data['G'])
        ws.cell(row=i, column=8, value=row_data['H'])
        ws.cell(row=i, column=15, value=row_data['O'])
        
        # 不锈钢密度 7.95，其他材质 7.85
        density_formula = f'IF(OR(D{i}="SUS",D{i}="不锈钢"),7.95,7.85)'
        ws.cell(row=i, column=9, value=f'=E{i}*F{i}*G{i}/1000000*{density_formula}')
        ws.cell(row=i, column=10, value=f'=I{i}*H{i}')
        ws.cell(row=i, column=11, value=f'=1-O{i}')
        ws.cell(row=i, column=12, value=f'=J{i}*(1-K{i})*0.85')
        ws.cell(row=i, column=14, value=f'=J{i}*M{i}')
    
    last_row = len(all_rows) + 2
    ws.cell(row=last_row, column=12, value=f'=SUM(L2:L{last_row-1})')
    ws.cell(row=last_row+1, column=12, value=2)
    ws.cell(row=last_row+2, column=12, value=f'=L{last_row}*L{last_row+1}')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    if output_path == 'auto':
        output_path = os.path.join(os.path.dirname(all_rows[0]['_pdf_path']), f'{order_no}_费用统计.xlsx')
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description='从 ByWork PDF 工单提取板材数据生成 Excel')
    parser.add_argument('--pdf', help='单个 PDF 文件路径')
    parser.add_argument('--batch', help='PDF 文件夹路径（批量处理）')
    parser.add_argument('--output', default='auto', help='输出 Excel 文件路径')
    parser.add_argument('--order', help='订单号（可选，默认从文件名提取）')
    parser.add_argument('--validate-only', action='store_true', help='只检验，不生成 Excel')
    
    args = parser.parse_args()
    
    if not args.pdf and not args.batch:
        print(json.dumps({'success': False, 'error': '请指定 --pdf 或 --batch 参数'}))
        sys.exit(1)
    
    all_rows = []
    all_errors = []
    all_warnings = []
    pdf_files = []
    
    if args.pdf:
        pdf_files = [args.pdf]
    elif args.batch:
        pdf_files = sorted([f for f in os.listdir(args.batch) if f.endswith('.pdf')])
        pdf_files = [os.path.join(args.batch, f) for f in pdf_files]
    
    for pdf_path in pdf_files:
        if not os.path.exists(pdf_path):
            all_errors.append(f'文件不存在：{pdf_path}')
            continue
        
        result = process_pdf(pdf_path, args.order)
        
        for row in result['rows']:
            row['_pdf_path'] = pdf_path
        
        all_rows.extend(result['rows'])
        all_errors.extend(result['errors'])
        all_warnings.extend(result['warnings'])
    
    if args.validate_only:
        print(json.dumps({
            'success': True,
            'validate_only': True,
            'pdf_count': len(pdf_files),
            'errors': all_errors,
            'warnings': all_warnings
        }, ensure_ascii=False))
        return
    
    if not all_rows:
        print(json.dumps({
            'success': False,
            'error': '未提取到任何数据',
            'errors': all_errors,
            'warnings': all_warnings
        }, ensure_ascii=False))
        return
    
    output_path = create_excel(all_rows, args.output, args.order or extract_order_from_filename(os.path.basename(pdf_files[0])))
    
    print(json.dumps({
        'success': True,
        'message': f'处理 {len(pdf_files)} 个 PDF 文件，生成 {len(all_rows)} 行数据',
        'output_file': output_path,
        'stats': {
            'pdf_count': len(pdf_files),
            'row_count': len(all_rows)
        },
        'errors': all_errors,
        'warnings': all_warnings
    }, ensure_ascii=False))


if __name__ == '__main__':
    main()
