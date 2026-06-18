#!/usr/bin/env python3
"""
PDF to Excel - 从 ByWork PDF 工单自动生成两种 Excel 表格
"""

import sys
import os
import argparse
import glob
import re
from datetime import datetime

sys.path.insert(0, '/home/admin/.local/lib/python3.11/site-packages')

try:
    import fitz
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    print(f'缺少依赖：{e}')
    sys.exit(1)


def extract_pdf_data(pdf_path):
    """从单个 PDF 提取所有 Sheet dimension 数据"""
    doc = fitz.open(pdf_path)
    results = []
    
    # 从第 1 页 Material data 提取厚度
    text = doc[0].get_text()
    thickness = 0
    for line in text.split('\n'):
        match = re.match(r'^(\d+\.?\d*)\s*mm$', line.strip())
        if match:
            thickness = float(match.group(1))
            break
    
    # 遍历所有页面提取 Sheet dimension、Cycles、Waste
    for page_num in range(len(doc)):
        text = doc[page_num].get_text()
        lines = text.split('\n')
        
        in_plan_data = False
        header_count = 0
        data_lines = []
        
        for line in lines:
            if 'Plan data' in line:
                in_plan_data = True
                continue
            # 遇到 Flat part data 或 Part No. 就停止
            if in_plan_data and ('Flat part data' in line or 'Part No.' in line):
                break
            
            if in_plan_data:
                # 表头可能跨行，如 "Number \nof \nparts"
                header_keywords = ['File name', 'Plan dimension', 'Sheet dimension', 
                                   'Cycles', 'Cutting time', 'Waste', 'Number', 'parts']
                if any(k in line for k in header_keywords):
                    header_count += 1
                    continue
                
                # 至少 7 个表头后开始收集数据
                if header_count >= 7 and line.strip():
                    data_lines.append(line)
        
        # 解析垂直排列的数据
        # 数据顺序：[Plan dimension, Sheet dimension, Cycles, Cutting time, Waste, Number of parts]
        if len(data_lines) >= 6:
            # 查找 Sheet dimension（第一个带 x 的尺寸）
            sheet_idx = -1
            for i, line in enumerate(data_lines):
                dim_match = re.search(r'(\d+)\s*x\s*(\d+)', line)
                if dim_match:
                    # 检查下一个数据是否是数字（Cycles）
                    if i+1 < len(data_lines) and data_lines[i+1].strip().isdigit():
                        sheet_idx = i
                        break
            
            if sheet_idx >= 0:
                sheet_dim_line = data_lines[sheet_idx]
                dim_match = re.search(r'(\d+)\s*x\s*(\d+)', sheet_dim_line)
                if dim_match:
                    length = int(dim_match.group(1))
                    width = int(dim_match.group(2))
                    
                    # Cycles 在 Sheet dimension 后面
                    quantity = 0
                    if sheet_idx+1 < len(data_lines) and data_lines[sheet_idx+1].strip().isdigit():
                        quantity = int(data_lines[sheet_idx+1])
                    
                    # Waste 在 Cutting time 后面（索引 +3）
                    waste_rate = 0
                    if sheet_idx+3 < len(data_lines):
                        waste_line = data_lines[sheet_idx+3]
                        waste_match = re.search(r'(\d+\.?\d*)\s*%', waste_line)
                        if waste_match:
                            waste_rate = float(waste_match.group(1)) / 100
                    
                    # Number of parts 在 Waste 后面（索引 +4）
                    number_of_parts = 0
                    if sheet_idx+4 < len(data_lines):
                        parts_line = data_lines[sheet_idx+4]
                        if parts_line.strip().isdigit():
                            number_of_parts = int(parts_line)
                    
                    results.append({
                        'thickness': thickness,
                        'length': length,
                        'width': width,
                        'quantity': quantity,  # 使用 Cycles（板材数量）
                        'number_of_parts': number_of_parts,  # 零件数量（用于验证）
                        'waste_rate': waste_rate
                    })
    
    return results


# ============================================================
# 板材价格配置（2026 年 5 月更新，支持从 Excel 自动读取当月价格）
# ============================================================
# 当提供 --prices 参数时，自动从 Excel 读取当月采购价格
# 未提供时使用下方硬编码价格作为 fallback
# ============================================================

MATERIAL_PRICES = {
    # Q235 开平板（减薄后厚度→元/kg）- 与 Excel 显示一致
    # 来源："26 年 6 月份板材价格" 表 col 37 "26年6月采购价格"
    # 标称厚度→减薄后：3→2.75, 4→3.75, 5→4.75, 6→5.75, 8→7.75, 10→9.75, 12→11.75
    'Q235': {
        2.5: 3.883, 2.75: 3.863, 3.75: 3.783,
        4.75: 3.743, 5.75: 3.743, 7.75: 3.743,
        9.75: 3.743, 11.75: 3.743
        # 13.75mm(14mm 标称) 价格表无数据，不填价格
    },
    # Q235 冷板（减薄后厚度→元/kg）
    'Q235 冷板': {
        1.0: 4.422, 1.5: 4.372, 2.0: 4.372,
        2.5: 4.602
    },
    # Q345 锰板（减薄后厚度→元/kg）
    # 标称厚度→减薄后：3→2.75, 4→3.75, 5→4.75, 6→5.75, 8→7.75, 10→9.75, 12→11.75
    'Q345': {
        2.75: 3.912, 3.75: 3.912, 4.75: 3.862,
        5.75: 3.761, 7.75: 3.761, 9.75: 3.761,
        11.75: 3.831
    },
}

# 特殊材质固定价格（元/kg）
SPECIAL_MATERIAL_PRICES = {
    '镀锌板': 5.0,    # 固定价
    'SUS': 16.0,      # 不锈钢
    '不锈钢': 16.0,   # 不锈钢
    '耐磨板': 6.0,    # 固定价
    # 花纹板价格留空，不填
}

def read_prices_from_excel(excel_path):
    """从板材价格 Excel 读取当月采购价格，返回 {材质: {减薄厚度: 单价}} 格式
    
    自动检测当月采购价格列（"XX月\n采购价格"），按材质和厚度提取价格。
    价格单位：元/吨 → 转换为 元/kg（÷1000）
    """
    try:
        import openpyxl as openpyxl_read
        wb = openpyxl_read.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f'⚠️ 无法读取价格文件 {excel_path}: {e}')
        return None
    
    # 查找当月价格表（"XX年X月份板材价格"）
    price_sheet = None
    max_month = 0
    for name in wb.sheetnames:
        if '板材价格' in name:
            # 提取月份："26年6月份板材价格" → 6
            m = re.search(r'(\d+)年(\d+)月', name)
            if m:
                year = int(m.group(1))
                month = int(m.group(2))
                # 取最大月份（当月）
                val = year * 100 + month
                if val > max_month:
                    max_month = val
                    price_sheet = name
    
    if not price_sheet:
        print(f'⚠️ 未找到板材价格表')
        return None
    
    ws = wb[price_sheet]
    print(f'📊 使用价格表: {price_sheet}')
    
    # 找采购价格列（"XX月\n采购价格" 或 "XX月\n采购价格"）
    price_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and isinstance(val, str) and '采购价格' in val:
            price_col = col
    
    if not price_col:
        print(f'⚠️ 未找到采购价格列')
        return None
    
    # 材质→减薄厚度→价格 的映射
    # Excel 中：Q235 开平板的厚度是减薄后的值（2.5, 3, 4, 4.75, 5.75, 7.75, 9.75, 11.75）
    prices = {}
    for row in range(3, ws.max_row + 1):
        name_val = ws.cell(row=row, column=3).value  # C列：名称
        thickness_val = ws.cell(row=row, column=6).value  # F列：厚度mm
        price_val = ws.cell(row=row, column=price_col).value  # 采购价格列
        
        if not name_val or not thickness_val or not price_val:
            continue
        
        name_str = str(name_val).strip()
        try:
            thickness = float(thickness_val)
            price_per_ton = float(price_val)
        except (ValueError, TypeError):
            continue
        
        # 跳过无效价格（0、负数、非数字）
        if price_per_ton <= 0 or price_per_ton > 100000:
            continue
        
        price_per_kg = round(price_per_ton / 1000, 3)
        
        # 判断材质分类
        if 'Q345' in name_str or '锰板' in name_str:
            material = 'Q345'
        elif '冷板' in name_str or 'ST12' in name_str:
            material = 'Q235 冷板'
        elif 'Q235' in name_str or '开平板' in name_str:
            material = 'Q235'
        else:
            continue
        
        # Q235/Q345 标称厚度≥3mm 时减薄 0.25mm（与代码逻辑一致）
        # Excel 中 4.75/5.75/7.75/9.75/11.75 已经是减薄后值，不需要再减
        # 只对整数厚度（如 3, 4, 5, 6, 8, 10, 12）执行减薄
        lookup_thickness = thickness
        if material in ('Q235', 'Q345') and thickness >= 3 and thickness == int(thickness):
            lookup_thickness = round(thickness - 0.25, 2)
        
        if material not in prices:
            prices[material] = {}
        prices[material][lookup_thickness] = price_per_kg
    
    # 打印读取结果
    for mat, p in sorted(prices.items()):
        items = ', '.join(f'{t}mm={v}' for t, v in sorted(p.items()))
        print(f'  {mat}: {items}')
    
    return prices


def get_price_per_kg(material, thickness, excel_prices=None):
    """根据材质和厚度获取单价（元/公斤），无价格返回 None
    
    优先使用 Excel 当月价格，fallback 到硬编码价格。
    """
    # 特殊材质固定价格
    if material in SPECIAL_MATERIAL_PRICES:
        return SPECIAL_MATERIAL_PRICES[material]
    
    # 优先使用 Excel 价格
    if excel_prices and material in excel_prices:
        prices = excel_prices[material]
        for t, price in prices.items():
            if abs(t - thickness) < 0.01:
                return price
        return None
    
    # fallback 到硬编码价格
    if material in MATERIAL_PRICES:
        prices = MATERIAL_PRICES[material]
        for t, price in prices.items():
            if abs(t - thickness) < 0.01:
                return price
        return None
    
    return None

def get_material_from_filename(pdf_file, thickness):
    """从文件名和厚度判断材质"""
    filename = os.path.basename(pdf_file).lower()
    if filename.startswith('sus-'):
        return 'SUS'
    elif filename.startswith('dxb-'):
        return '镀锌板'
    elif filename.startswith('hwb-') or 'hwb' in filename.lower():  # hwb = 花纹板
        return '花纹板'
    elif 'mn' in filename.lower():  # Mn 文件是 Q345
        return 'Q345'
    elif thickness <= 2:
        return 'Q235 冷板'
    else:
        return 'Q235'


def create_material_cost_table(pdf_dir, output_file, contract_no, excel_prices=None):
    """生成板材费用计算表"""
    pdf_files = sorted(glob.glob(os.path.join(pdf_dir, '*.pdf')))
    
    # 提取所有 PDF 数据
    data_rows = []
    errors = []  # 收集错误
    
    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_dir, pdf_file)
        results = extract_pdf_data(pdf_path)
        
        material = get_material_from_filename(pdf_file, results[0]['thickness'] if results else 0)
        
        for result in results:
            # 检查零件数量是否为 0
            if result.get('number_of_parts', 0) == 0:
                error_msg = f'{pdf_file} - Sheet {result["length"]}x{result["width"]}mm 的零件数量为 0'
                errors.append(error_msg)
            
            data_rows.append({
                'material': material,
                'length': result['length'],
                'width': result['width'],
                'thickness': result['thickness'],
                'quantity': result['quantity'],
                'waste_rate': result['waste_rate'],
                'number_of_parts': result.get('number_of_parts', 0)
            })
    
    # 如果有错误，报错并退出
    if errors:
        print('\n❌ **错误：以下板材的零件数量为 0，请重新生成 PDF！**')
        for err in errors:
            print(f'  - {err}')
        print('\n已终止生成，请修正后重试。')
        sys.exit(1)
    
    # 排序：材质（Q235 冷板 → Q235 → Q345 → 其他），厚度从低到高
    material_order = {'Q235 冷板': 0, 'Q235': 1, 'Q345': 2}
    data_rows.sort(key=lambda x: (material_order.get(x['material'], 3), x['thickness']))
    
    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '板材费用统计'
    
    headers = ['序号', '日期', '订单号', '材质（Q235/Q345/SUS）', '长（mm）', '宽（mm）',
               '厚（mm）', '数量', '重量（单重 Kg）', '合计重量（Kg）', '板材利用率',
               '废料重量（Kg）', '单价（Kg）', '板材价格', '废料率%']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='000000', size=11)  # 黑色字体
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # 取消背景色
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    for i, row_data in enumerate(data_rows, 2):
        # 序号
        ws.cell(row=i, column=1, value=i-1)
        # 日期（留空，用户自己填）
        # 订单号（留空，用户自己填）
        # 材质
        ws.cell(row=i, column=4, value=row_data['material'])
        # 长
        ws.cell(row=i, column=5, value=row_data['length'])
        # 宽
        ws.cell(row=i, column=6, value=row_data['width'])
        # 厚（实际厚度：Q235/Q345 ≥3mm 时减 0.25mm）
        thickness = row_data['thickness']
        material = row_data['material']
        display_thickness = thickness
        if material in ['Q235', 'Q235 冷板', 'Q345'] and thickness >= 3:
            display_thickness = round(thickness - 0.25, 2)
        ws.cell(row=i, column=7, value=display_thickness)
        # 数量
        ws.cell(row=i, column=8, value=row_data['quantity'])
        # 废料率
        ws.cell(row=i, column=15, value=row_data['waste_rate'])
        
        # 单价（元/公斤）- 根据材质和减薄后厚度查找
        price_per_kg = get_price_per_kg(material, display_thickness, excel_prices)
        if price_per_kg:
            ws.cell(row=i, column=13, value=round(price_per_kg, 3))
        
        density_formula = f'IF(OR(D{i}="SUS",D{i}="不锈钢"),7.95,7.85)'
        ws.cell(row=i, column=9, value=f'=ROUND(E{i}*F{i}*G{i}/1000000*{density_formula},2)')
        ws.cell(row=i, column=10, value=f'=ROUND(I{i}*H{i},2)')
        ws.cell(row=i, column=11, value=f'=1-O{i}')
        ws.cell(row=i, column=12, value=f'=ROUND(J{i}*(1-K{i})*0.85,2)')
        ws.cell(row=i, column=14, value=f'=ROUND(J{i}*M{i},2)')
        
        for col in range(1, 16):
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=10)
            cell.border = thin_border
    
    # 汇总行起始位置
    summary_start = len(data_rows) + 2
    
    # 板材价格合计（合并 A-K 列，居中）
    ws.merge_cells(f'A{summary_start}:K{summary_start}')
    ws.cell(row=summary_start, column=1, value='板材价格合计')
    ws.cell(row=summary_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=summary_start, column=14, value=f'=SUM(N2:N{summary_start-1})')
    for col in range(1, 16):
        ws.cell(row=summary_start, column=col).border = thin_border
    
    # 废料重量（合并 A-K 列，居中，无填充）
    ws.merge_cells(f'A{summary_start+1}:K{summary_start+1}')
    ws.cell(row=summary_start+1, column=1, value='废料重量')
    ws.cell(row=summary_start+1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=summary_start+1, column=12, value=f'=SUM(L2:L{summary_start-1})')
    for col in range(1, 16):
        ws.cell(row=summary_start+1, column=col).border = thin_border
    
    # 废料单价（合并 A-K 列，居中，无填充）
    ws.merge_cells(f'A{summary_start+2}:K{summary_start+2}')
    ws.cell(row=summary_start+2, column=1, value='废料单价')
    ws.cell(row=summary_start+2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=summary_start+2, column=12, value=2)
    for col in range(1, 16):
        ws.cell(row=summary_start+2, column=col).border = thin_border
    
    # 废料价格（合并 A-K 列，居中，无填充）
    ws.merge_cells(f'A{summary_start+3}:K{summary_start+3}')
    ws.cell(row=summary_start+3, column=1, value='废料价格')
    ws.cell(row=summary_start+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=summary_start+3, column=12, value=f'=L{summary_start+1}*L{summary_start+2}')
    for col in range(1, 16):
        ws.cell(row=summary_start+3, column=col).border = thin_border
    
    # 设置这四行的 L、N 和 O 列也居中
    for row_offset in range(4):
        for col in [12, 14, 15]:
            ws.cell(row=summary_start+row_offset, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    column_widths = {'A': 6, 'B': 12, 'C': 16, 'D': 14, 'E': 10, 'F': 10, 'G': 10, 'H': 8, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 10, 'N': 12, 'O': 10}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 第一行行高调大
    ws.row_dimensions[1].height = 60
    for i in range(2, summary_start + 5):
        ws.row_dimensions[i].height = 20
    
    for i in range(2, summary_start):
        # 长、宽、数量为整数
        for col in [5, 6, 8]:
            ws.cell(row=i, column=col).number_format = '0'
        # 厚度保留 2 位小数（显示为 2.00 而不是 2，2.75 不四舍五入）
        ws.cell(row=i, column=7).number_format = '0.00'
        # 重量、价格为 2 位小数
        for col in [9, 10, 12, 14]:
            ws.cell(row=i, column=col).number_format = '0.00'
        # 利用率、废料率为百分比 2 位小数
        ws.cell(row=i, column=11).number_format = '0.00%'
        ws.cell(row=i, column=15).number_format = '0.00%'
    
    wb.save(output_file)
    return output_file


def extract_cutting_times_from_pdf(pdf_path):
    """从 PDF 第一页提取 Cutting time"""
    doc = fitz.open(pdf_path)
    text = doc[0].get_text()
    lines = text.split('\n')
    
    thickness = None
    in_material = False
    for line in lines:
        if 'Material data' in line:
            in_material = True
            continue
        if in_material and 'Plan data' in line:
            break
        if in_material:
            match = re.match(r'^(\d+\.?\d*)\s*mm$', line.strip())
            if match:
                thickness = float(match.group(1))
                break
    
    cutting_time = None
    for i, line in enumerate(lines):
        if 'Cutting time' in line:
            for j in range(i+1, min(i+15, len(lines))):
                match = re.search(r'(\d+\.?\d*)\s*min', lines[j])
                if match:
                    cutting_time = float(match.group(1))
                    break
            if cutting_time:
                break
    
    if cutting_time and thickness:
        return [(thickness, cutting_time)]
    return []


def extract_cutting_times_batch(pdf_dir):
    """从目录中所有 PDF 提取 Cutting time 并累加"""
    pdf_files = glob.glob(os.path.join(pdf_dir, '*.pdf'))
    thickness_times = {}
    
    for pdf_path in pdf_files:
        result = extract_cutting_times_from_pdf(pdf_path)
        if result:
            thickness, cutting_time = result[0]
            if thickness in thickness_times:
                thickness_times[thickness] += cutting_time
            else:
                thickness_times[thickness] = cutting_time
    
    return thickness_times


def create_zhenyuan_settlement(pdf_dir, output_file, contract_no):
    """生成震源机械结算单"""
    thickness_times = extract_cutting_times_batch(pdf_dir)
    
    data_rows = []
    for thickness in sorted(thickness_times.keys()):
        cutting_time = thickness_times[thickness]
        data_rows.append({
            'date': '',
            'thickness': thickness,
            'theoretical_time': round(cutting_time, 2),
            'processing_fee': 6.33
        })
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = contract_no
    
    company_info = {
        'name_cn': '南通震源机械有限公司',
        'name_en': 'Nantong source of metal cutting Co.Ltd.',
        'address': '地址：江苏省如皋市中山西路 398 号',
        'contact': '联系人：谢万宏   电话：18962729888   传真：86-0513-87872988',
        'web': 'Http://www.ntzhenyuan.com   E-mail：2995279201@qq.com'
    }
    
    normal_font = Font(size=10, name='宋体')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    summary_font = Font(bold=True, size=10, name='宋体')
    summary_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # 公司抬头
    ws.cell(row=1, column=1, value=company_info['name_cn'])
    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1).font = Font(bold=True, size=24, name='宋体')
    ws.cell(row=1, column=1).alignment = center_align
    
    ws.cell(row=2, column=1, value=company_info['name_en'])
    ws.merge_cells('A2:L2')
    ws.cell(row=2, column=1).alignment = center_align
    
    ws.cell(row=3, column=1, value=company_info['address'])
    ws.merge_cells('A3:L3')
    ws.cell(row=3, column=1).alignment = center_align
    
    ws.cell(row=4, column=1, value=company_info['contact'])
    ws.merge_cells('A4:L4')
    ws.cell(row=4, column=1).alignment = center_align
    
    ws.cell(row=5, column=1, value=company_info['web'])
    ws.merge_cells('A5:L5')
    ws.cell(row=5, column=1).alignment = center_align
    
    ws.row_dimensions[6].height = 15
    
    # 合同编号
    contracts = [
        f'1.合同编号：{contract_no}',
        '2.合同编号：AW-26-0215',
        '3.合同编号：AW-26-0229',
        '4.合同编号：AW-26-0278',
        '5.合同编号：AW-26-0339'
    ]
    
    for i, contract in enumerate(contracts, 7):
        ws.cell(row=i, column=1, value=contract)
        ws.merge_cells(f'A{i}:L{i}')
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    ws.row_dimensions[12].height = 14.25
    
    # 表头
    headers = [
        '生产日期', '厚度 mm', '', '理论时间分钟', '理论时间\n总计分钟',
        '结算时间分钟', '加工费\n1 分钟', '加工费 总价（元）',
        '材料费', '打孔攻丝总价', '折弯总价', '原材料 + 打孔攻丝费 + 折弯\n总计元'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    for row in range(1, 14):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = thin_border
    
    # 结算系数
    for col in range(1, 13):
        cell = ws.cell(row=14, column=col)
        cell.border = thin_border
        cell.alignment = center_align
    ws.cell(row=14, column=6, value=0.1)
    ws.cell(row=14, column=6).number_format = '0.0'
    
    # 数据行（只填充有数据的行，不留空白行）
    standard_thickness = [1, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 14, 16, 18, 20]
    
    row = 15
    for thickness in standard_thickness:
        row_data = None
        
        for dr in data_rows:
            if abs(dr['thickness'] - thickness) < 0.01:
                row_data = dr
                break
        
        if row_data:
            ws.cell(row=row, column=2, value=thickness)
            ws.cell(row=row, column=5, value=row_data.get('theoretical_time', 0))
            ws.cell(row=row, column=6, value=f'=E{row}+(E{row}*F$14)')
            ws.cell(row=row, column=7, value=row_data.get('processing_fee', 0))
            ws.cell(row=row, column=8, value=f'=G{row}*F{row}')
            ws.cell(row=row, column=12, value=f'=H{row}+I{row}+J{row}+K{row}')
            
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align
            
            row += 1  # 只有有数据时才增加行号
    
    # 合计行（在最后一个数据行之后）
    summary_row = row  # row 是上一个数据行 +1
    ws.cell(row=summary_row, column=1, value='合计：')
    ws.cell(row=summary_row, column=1).font = summary_font
    ws.cell(row=summary_row, column=1).fill = summary_fill
    ws.cell(row=summary_row, column=12, value=f'=SUM(L15:L{summary_row-1})')
    ws.cell(row=summary_row, column=12).number_format = '0.00'
    ws.cell(row=summary_row, column=12).font = summary_font
    
    for col in range(1, 13):
        cell = ws.cell(row=summary_row, column=col)
        cell.border = thin_border
        cell.alignment = center_align
    
    # 列宽和行高
    column_widths = {'A': 11.0, 'B': 13.0, 'C': 13.0, 'D': 13.0, 'E': 13.0, 'F': 13.0, 'G': 13.0, 'H': 12.625, 'I': 13.0, 'J': 13.0, 'K': 13.0, 'L': 12.625}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[1].height = 31.5
    ws.row_dimensions[2].height = 20.25
    for row in range(3, 6):
        ws.row_dimensions[row].height = 15.0
    for row in range(7, 12):
        ws.row_dimensions[row].height = 15.0
    ws.row_dimensions[12].height = 14.25
    ws.row_dimensions[13].height = 57.0
    ws.row_dimensions[14].height = 15.0
    # 数据行行高（动态）
    for row in range(15, summary_row):
        ws.row_dimensions[row].height = 14.25
    # 合计行行高
    ws.row_dimensions[summary_row].height = 18.75
    
    # 数字格式（动态）
    for row in range(15, summary_row):
        thickness = ws.cell(row=row, column=2).value
        if thickness and thickness == int(thickness):
            ws.cell(row=row, column=2).number_format = '0'
        else:
            ws.cell(row=row, column=2).number_format = '0.0'
        ws.cell(row=row, column=5).number_format = '0.00'
        ws.cell(row=row, column=6).number_format = '0.00'
        ws.cell(row=row, column=7).number_format = '0.00'
        ws.cell(row=row, column=8).number_format = '0.00'
        ws.cell(row=row, column=12).number_format = '0.00'
    
    wb.save(output_file)
    return output_file


def main():
    parser = argparse.ArgumentParser(description='从 PDF 工单生成 Excel')
    parser.add_argument('--pdf-dir', required=True, help='PDF 文件夹')
    parser.add_argument('--output', help='输出文件路径')
    parser.add_argument('--contract', default=None, help='合同编号（默认从 pdf-dir 路径自动提取）')
    parser.add_argument('--type', choices=['material', 'settlement', 'both'], default='both')
    parser.add_argument('--prices', help='板材价格 Excel 文件路径（自动读取当月采购价格）')
    
    args = parser.parse_args()
    
    # Auto-detect contract number from PDF directory name if not specified
    if not args.contract:
        import os
        dir_name = os.path.basename(args.pdf_dir.rstrip('/'))
        # Match pattern like AW-26-0994
        import re
        match = re.search(r'(AW-\d{2}-\d{4})', dir_name)
        if match:
            args.contract = match.group(1)
        else:
            args.contract = dir_name
    
    # 从 Excel 读取当月价格
    # 优先级：--prices 参数 > prices/ 目录下最新文件 > 硬编码 fallback
    excel_prices = None
    prices_source = None
    
    if args.prices:
        # 用户指定了价格文件
        prices_source = args.prices
    else:
        # 自动查找 prices/ 目录下最新的价格文件
        prices_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'prices')
        if os.path.isdir(prices_dir):
            xlsx_files = glob.glob(os.path.join(prices_dir, '*.xlsx')) + glob.glob(os.path.join(prices_dir, '*.xls'))
            if xlsx_files:
                # 取最新修改时间的文件
                prices_source = max(xlsx_files, key=os.path.getmtime)
    
    if prices_source:
        excel_prices = read_prices_from_excel(prices_source)
        if excel_prices:
            print(f'✅ 已从 {prices_source} 读取当月价格')
        else:
            print(f'⚠️ 价格读取失败（{prices_source}），使用默认价格')
    
    if args.type in ['material', 'both']:
        if args.output:
            out_dir = os.path.dirname(args.output) or '.'
            output_material = os.path.join(out_dir, f'{args.contract}_板材费用表.xlsx')
        else:
            output_material = f'/tmp/{args.contract}_板材费用表.xlsx'
        print(f'正在生成板材费用表...')
        create_material_cost_table(args.pdf_dir, output_material, args.contract, excel_prices)
        print(f'✅ 板材费用表已生成：{output_material}')
    
    if args.type in ['settlement', 'both']:
        if args.output:
            out_dir = os.path.dirname(args.output) or '.'
            output_settlement = os.path.join(out_dir, f'{args.contract}_震源结算单.xlsx')
        else:
            output_settlement = f'/tmp/{args.contract}_震源结算单.xlsx'
        print(f'正在生成震源结算单...')
        create_zhenyuan_settlement(args.pdf_dir, output_settlement, args.contract)
        print(f'✅ 震源结算单已生成：{output_settlement}')
    
    print(f'\n✅ 处理完成！')


if __name__ == '__main__':
    main()
