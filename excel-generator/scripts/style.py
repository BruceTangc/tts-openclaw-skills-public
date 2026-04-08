#!/usr/bin/env python3
"""
Excel 样式脚本 - 完整样式支持
"""

import sys
import json
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def apply_styles(file_path, style_config):
    """
    应用样式到 Excel 文件
    
    Args:
        file_path: Excel 文件路径
        style_config: 样式配置
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 应用标题行样式
        if style_config.get('header'):
            apply_header_style(ws, style_config['header'])
        
        # 应用边框
        if style_config.get('border'):
            apply_border(ws)
        
        # 自动列宽
        if style_config.get('auto_width'):
            auto_adjust_column_widths(ws)
        
        # 交替行
        if style_config.get('alternating_rows'):
            apply_alternating_rows(ws)
        
        # 保存
        wb.save(file_path)
        
        return {
            'success': True,
            'output_file': file_path,
            'styles_applied': list(style_config.keys())
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def apply_header_style(ws, header_config):
    """应用标题行样式"""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        
        # 加粗
        if header_config.get('bold'):
            cell.font = Font(bold=True)
        
        # 背景色
        if header_config.get('bgColor'):
            cell.fill = PatternFill(
                start_color=header_config['bgColor'],
                end_color=header_config['bgColor'],
                fill_type='solid'
            )
        
        # 文字颜色
        if header_config.get('color'):
            cell.font = Font(
                color=header_config['color'],
                bold=cell.font.bold if hasattr(cell.font, 'bold') else False
            )
        
        # 字体大小
        if header_config.get('fontSize'):
            cell.font = Font(
                size=header_config['fontSize'],
                bold=cell.font.bold if hasattr(cell.font, 'bold') else False,
                color=cell.font.color.rgb if hasattr(cell.font, 'color') and cell.font.color else None
            )
        
        # 对齐
        cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_border(ws):
    """应用边框到所有单元格"""
    thin_border = Side(style='thin', color='000000')
    full_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = full_border

def auto_adjust_column_widths(ws):
    """自动调整列宽"""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def apply_alternating_rows(ws):
    """应用交替行样式"""
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                    cell.fill = light_fill

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='应用样式到 Excel')
    parser.add_argument('--file', required=True, help='Excel 文件路径')
    parser.add_argument('--style', required=True, help='JSON 格式样式配置')
    
    args = parser.parse_args()
    style_config = json.loads(args.style)
    
    result = apply_styles(args.file, style_config)
    print(json.dumps(result, ensure_ascii=False))
