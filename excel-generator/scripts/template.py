#!/usr/bin/env python3
"""
Excel 模板填充脚本
"""

import sys
import json
import argparse
from openpyxl import load_workbook

def fill_template(template_path, output_path, data, replacements):
    """
    填充 Excel 模板
    
    Args:
        template_path: 模板文件路径
        output_path: 输出文件路径
        data: 要填充的数据
        replacements: 替换规则 {placeholder: value}
    """
    try:
        # 加载模板
        wb = load_workbook(template_path)
        ws = wb.active
        
        # 替换占位符
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    for placeholder, value in replacements.items():
                        cell.value = cell.value.replace(placeholder, str(value))
        
        # 追加数据（如果有）
        if data:
            for row_data in data:
                ws.append(row_data)
        
        # 保存
        wb.save(output_path)
        
        return {
            'success': True,
            'output_file': output_path,
            'replacements_count': len(replacements),
            'data_rows': len(data) if data else 0
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='填充 Excel 模板')
    parser.add_argument('--template', required=True, help='模板文件路径')
    parser.add_argument('--output', required=True, help='输出文件路径')
    parser.add_argument('--data', help='JSON 格式数据数组')
    parser.add_argument('--replacements', help='JSON 格式替换规则')
    
    args = parser.parse_args()
    
    data = json.loads(args.data) if args.data else []
    replacements = json.loads(args.replacements) if args.replacements else {}
    
    result = fill_template(args.template, args.output, data, replacements)
    print(json.dumps(result, ensure_ascii=False))
